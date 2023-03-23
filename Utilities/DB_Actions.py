import sqlite3
import openpyxl
from datetime import datetime
import json
import os
import pandas as pd
import Utilities.config as config
import getpass

def create_table(conn):
    """connect to database"""
    _create_db_table(conn)


def create_db_auth_table(conn):
    """Create Auth table with base passcodes"""
    _create_db_auth_table(conn)
    _delete_existing_rows(conn)
    _create_base_auth_entries(conn)

def delete_scripts_table(conn):
    """delete scripts table"""
    _delete_scripts_table(conn)
    
def delete_existing_tool_rows(conn):
    """delete existing tool rows"""
    _delete_existing_tool_rows(conn)
        
def update_passcode(conn, passcode_dict):
    """Update passcode"""
    return _update_passcode(conn, passcode_dict)

def validate_passcode(conn, passcode_dict):
    """Authenticate passcode"""
    return _validate_passcode(conn, passcode_dict)

def import_data(conn, input_file):
    """Insert data into DB from Excel file"""
    _import_data(conn, input_file)


def insert_data(conn, data_dict):
    """Insert new row into ScriptInfo table"""
    return _insert_data(conn, data_dict)


def update_data(conn, data, _id):
    """Update row in ScriptInfo table"""
    _update_data(conn, data, _id)


def delete_data(conn, _id):
    """Delete row from ScriptInfo table"""
    _delete_data(conn, _id)

def undelete_data(conn, _id):
    """UnDelete row from ScriptInfo table"""
    _undelete_data(conn, _id)

def get_data(conn, search_dict=None, export_as='json'):
    """Get data from DB"""
    return _get_data(conn, search_dict, export_as)


def get_unique_values(conn, column, search_dict=None):
    """Get list of values from selected column"""
    unique_list = _get_unique_values(conn, column, search_dict)
    return unique_list


def get_db_connection(db_path):
    """connect to database (create a new one if it doesn't exist)"""
    return sqlite3.connect(db_path)


def _import_data(conn, input_file):
    """Insert data into DB from Excel file"""
    valid_columns = ['team', 'title', 'description', 'software',
                     'keyword', 'function', 'owner', 'toolpath', 'documentpath']
    """Open input file"""
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active

    """Get the header row and create a mapping of column names to indices"""
    header_row = list(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    column_mapping = {header_row[i].lower().strip(): i for i in range(len(header_row))}

    """Get the data rows and insert them into the database"""
    for row in worksheet.iter_rows(min_row=3, values_only=True):
        columns = []
        values = []
        has_valid_columns = False
        for column_name in valid_columns:
            """If the column name is not in the input file, skip it"""
            if column_name not in column_mapping:
                continue
            column_index = column_mapping[column_name]
            value = row[column_index]
            """Comma separate Keyword by + ad / """
            if column_name.lower() == 'keyword' and (isinstance(value, str) and ('+' in value or '/' in value)):
                value = str(','.join(value.split('+')).replace('/', ',')).strip().replace(" ,", ",")
            columns.append(column_name)
            values.append(value)
            has_valid_columns = True

        """Skip the row if there are no valid columns"""
        if not has_valid_columns:
            continue

        """Add additional default columns and values"""
        columns.append('Status')
        columns.append('Created_On')
        columns.append('Created_By')
        columns.append('Version')
        
        values.append('Active')
        values.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        values.append(getpass.getuser())
        values.append(1)
                
        """Construct the INSERT INTO query"""
        columns = ','.join(columns)
        query = f"INSERT INTO ScriptInfo ({columns}) " \
                f"VALUES ({','.join(['?'] * (len(columns.split(','))))})"

        """Insert the row into the database"""
        conn.execute(query, values)

    """Commit the changes to the database"""
    conn.commit()


def _create_db_table(conn):
    """Create table if it doesn't already exist"""
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='ScriptInfo'")
    result = cursor.fetchone()
    if result is None:
        cursor.execute('''
            CREATE TABLE ScriptInfo (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Team TEXT,
                Title TEXT,
                Description TEXT,
                Software TEXT,
                Keyword TEXT,
                Function TEXT,
                Owner TEXT,
                ToolPath TEXT,
                Status TEXT,
                DocumentPath TEXT,
                Created_On DATETIME,
                Edited_On DATETIME,
                Deleted_On DATETIME,
                Created_By TEXT,
                Edited_By TEXT,
                Deleted_By TEXT,
                Version INTEGER
            );
        ''')
    conn.commit()


def _create_db_auth_table(conn):
    """Create table if it doesn't already exist"""
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='Auth'")
    result = cursor.fetchone()
    if result is None:
        cursor.execute('''
            CREATE TABLE Auth (
                ID INTEGER PRIMARY KEY AUTOINCREMENT,
                Passcode TEXT,
                AuthLevel TEXT
            );
        ''')
    conn.commit()

def _delete_scripts_table(conn):
    """Delete the ScriptInfo table if it exists"""
    cursor = conn.cursor()
    cursor.execute("DROP TABLE IF EXISTS ScriptInfo;")
    conn.commit()
    
def _delete_existing_tool_rows(conn):
    """Delete existing rows in ScriptInfo table if they exist"""
    cursor = conn.cursor()
    cursor.execute("DELETE FROM ScriptInfo;")
    conn.commit()
    
def _delete_existing_rows(conn):
    """Delete existing rows in Auth table if they exist"""
    cursor = conn.cursor()
    cursor.execute("DELETE FROM Auth;")
    conn.commit()

def _create_base_auth_entries(conn):
    """Create a dictionary with two entries, and insert them into the Auth table"""
    auth_dict = {"passcode": 'Super#123', "authlevel": "Super"}, {"passcode": 'Admin@123', "authlevel": "Admin"}
    cursor = conn.cursor()
    for auth in auth_dict:
        cursor.execute("INSERT INTO Auth (Passcode, AuthLevel) VALUES (?, ?);", (auth['passcode'], auth['authlevel']))
    conn.commit()


def _update_passcode(conn, passcode_dict):
    """Update passcode in Auth table if old and new passcodes are valid"""
    old_passcode = passcode_dict['old_passcode']
    new_passcode = passcode_dict['new_passcode']
    cursor = conn.cursor()
    
    """Check if new passcodes is not the same as the passcode with authlevel 'Super'"""
    cursor.execute("SELECT Passcode FROM Auth WHERE AuthLevel = 'Super';")
    result = cursor.fetchone()
    if result is None:
        return "No passcode with authlevel 'Super' found in Auth table"
    super_passcode = result[0]
    
    cursor.execute("SELECT Passcode FROM Auth WHERE AuthLevel = 'Admin';")
    result = cursor.fetchone()
    admin_passcode = result[0]
    if old_passcode != super_passcode and old_passcode != admin_passcode:
        return "Invalid passcode"
    if new_passcode == super_passcode:
        return "Forbidden passcode"
    """Replace passcode as new_passcode where authlevel = 'Admin' in Auth table"""
    cursor.execute(f"UPDATE Auth SET Passcode = '{new_passcode}' WHERE authlevel = 'Admin';")
    conn.commit()
    return "Passcode updated successfully"


def _validate_passcode(conn, passcode_dict):
    """Validate passcode and get authlevel from Auth table"""
    passcode = passcode_dict['passcode']
    cursor = conn.cursor()
    """Check if passcode exists in Auth table"""
    cursor.execute("SELECT AuthLevel FROM Auth WHERE Passcode = ?;", (passcode,))
    result = cursor.fetchone()
    if result is None:
        return "Invalid passcode"
    authlevel = result[0]
    return authlevel

def _check_for_duplicate(conn, data_dict1, _id=None):
    """Check if data already exists in ScriptInfo table"""
    if _id:
        data_dict1.update({'ID': _id})
    if 'Version' in data_dict1:
        del data_dict1['Version']
    cursor = conn.cursor()
    placeholders = ' AND '.join([f"{key} = ?" for key in data_dict1.keys()])
    sql = f'''SELECT * FROM ScriptInfo WHERE {placeholders}'''
    values = tuple(data_dict1.values())
    cursor.execute(sql, values)
    result = cursor.fetchone()
    if result is not None:
        print("Data already exists in ScriptInfo table")
        return True
    else:
        return False


def _insert_data(conn, data_dict):
    """Check if data already exists in ScriptInfo table"""
    data_exists_in_ScriptInfo = _check_for_duplicate(conn, data_dict)
    if not data_exists_in_ScriptInfo:
        """Insert new row into ScriptInfo table"""
        cursor = conn.cursor()
        
        if 'Version' not in data_dict:
            data_dict.update({'Version': 1})
        data_dict.update({'Created_On': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
        data_dict.update({'Created_by': getpass.getuser()})
        placeholders = ', '.join(['?' for _ in range(len(data_dict))])
        columns = ', '.join(data_dict.keys())
        values = tuple(data_dict.values())
        sql = f"INSERT INTO ScriptInfo ({columns}) VALUES ({placeholders})"
        cursor.execute(sql, values)
        conn.commit()
        msg = 'Record inserted in tool database'
        inserted = True
    else:
        msg = "Data already exists in ScriptInfo table"
        inserted = False
    return {'inserted': inserted, 'message': msg}

def _update_data(conn, data_dict, _id):
    """Check if data already exists in ScriptInfo table"""
    data_exists_in_ScriptInfo = _check_for_duplicate(conn, data_dict, _id)
    if not data_exists_in_ScriptInfo:
        """Update row in ScriptInfo table"""
        cursor = conn.cursor()
        if not (len(data_dict) == 2 and ('ToolPath' in data_dict or 'DocumentPath' in data_dict)):
            data_dict.update({'Edited_On': datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
            data_dict.update({'Edited_By': getpass.getuser()})
        placeholders = ', '.join([f"{key} = ?" for key in data_dict])
        values = tuple(data_dict.values())
        sql = f"UPDATE ScriptInfo SET {placeholders} WHERE ID = ?"
        cursor.execute(sql, (*values, _id))
        conn.commit()


def _delete_data(conn, _id):
    """Check if data already exists in ScriptInfo table"""
    data_dict = {'Status': 'Deleted'}
    data_deleted_in_ScriptInfo = _check_for_duplicate(conn, data_dict, _id)
    if not data_deleted_in_ScriptInfo:
        """Delete row from ScriptInfo table"""
        cursor = conn.cursor()
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute(f"UPDATE ScriptInfo SET Status='Deleted', Deleted_On='{current_time}', Deleted_By = '{getpass.getuser()}' WHERE ID={_id}")
        conn.commit()

def _undelete_data(conn, _id):
    """Check if data already exists in ScriptInfo table"""
    data_dict = {'Status': 'Active'}
    data_deleted_in_ScriptInfo = _check_for_duplicate(conn, data_dict, _id)
    if not data_deleted_in_ScriptInfo:
        """Delete row from ScriptInfo table"""
        cursor = conn.cursor()
        cursor.execute(f"UPDATE ScriptInfo SET Status='Active' WHERE ID={_id}")
        conn.commit()

def _get_data(conn, search_dict=None, export_as='json'):
    """get column names"""
    cur = conn.cursor()
    cur.execute("PRAGMA table_info('ScriptInfo')")
    columns = cur.fetchall()
    column_names = [column[1] for column in columns if column]

    """get entries from db"""
    if search_dict:
        
        sw_filter = search_dict['software'] if 'software' in search_dict else None
        fn_filter = search_dict['function'] if 'function' in search_dict else None
        kw_filter = search_dict['keyword'] if 'keyword' in search_dict else None
        wildcard = search_dict['wildcard'] if 'wildcard' in search_dict else None
        active_only = search_dict['ActiveRecordsOnly'] if 'ActiveRecordsOnly' in search_dict else True
        if active_only:
            query = "SELECT * FROM ScriptInfo WHERE Status = 'Active' "
        else:
            query = "SELECT * FROM ScriptInfo WHERE Status not null "
        
        if wildcard and len(wildcard) > 0:
            wild_query = f"{' OR '.join([f'{col} LIKE ?' for col in column_names])}"
            query += ' and ' + wild_query
            params = ['%' + wildcard + '%' for i in range(len(column_names))]
            cur.execute(query, params)
        else:
            if sw_filter and len(sw_filter) > 0 and sw_filter != 'All':
                query += f"and software = '{search_dict['software']}'"
            if fn_filter and len(fn_filter) > 0 and fn_filter != 'All':
                query += f"and function = '{search_dict['function']}'"
            if kw_filter and len(kw_filter) > 0 and kw_filter != 'All':
                query += f"and keyword like '%{search_dict['keyword']}%'"
            cur.execute(query)
    else:
        query = "SELECT * FROM ScriptInfo"
        cur.execute(query)

    """get data from db"""
    rows = cur.fetchall()
    
    if export_as=='json':
        """get all results and convert to JSON"""
        results = []
        for row in rows:
            result = {}
            for i in range(len(column_names)):
                result[column_names[i]] = row[i]
            results.append(result)
    elif export_as=='df':
        """get all results and convert to dataframe"""
        results = pd.DataFrame(rows, columns=column_names)
    else:
        results = rows
    return results


def _get_unique_values(conn, column, search_dict=None):
    """Get unique values for a given column and concatenate them with comma separation"""
    if search_dict:
        query = f"SELECT GROUP_CONCAT(DISTINCT {column}) FROM ScriptInfo " \
                "WHERE Status = 'Active' "
        if 'software' in search_dict and search_dict['software'] != 'All':
            query += f"and software = '{search_dict['software']}'"
        if 'function' in search_dict and search_dict['function'] != 'All':
            query += f"and function = '{search_dict['function']}'"
        if 'keyword' in search_dict and search_dict['keyword'] != 'All':
            query += f"and keyword like '%{search_dict['keyword']}%'"
        
    else:
        query = f"SELECT GROUP_CONCAT(DISTINCT {column}) FROM ScriptInfo WHERE Status = 'Active'"
    db_result = conn.execute(query).fetchone()[0]
    if db_result:
        result_set = set(db_result.strip().replace(" ,", ",").replace(", ", ",").split(','))
        result_sorted = sorted(list(result_set))
        
        result = ['All']
        for item in result_sorted:
            result.append(item)
        
    else:
        result = []
    return result


def main():
    Repo_Path = config.Repo_Path
    db_path = os.path.join(os.path.join(Repo_Path, 'Database'), 'SST.db')
    print('db_path', db_path)
    conn = get_db_connection(db_path)
    create_db_auth_table(conn)
    
    passcode_dict = {'passcode': 'Admin@123s'}
    authlevel = validate_passcode(conn, passcode_dict)
    print(1, authlevel)
    passcode_dict = {'passcode': 'Super#123'}
    authlevel = validate_passcode(conn, passcode_dict)
    print(2, authlevel)
    
    
    passcode_dict = {"old_passcode": 'Super#123', "new_passcode": 'Super#1123'}
    result = update_passcode(conn, passcode_dict)
    print(result)
    


if __name__ == '__main__':
    main()
